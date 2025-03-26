import pandas as pd
from openpyxl import load_workbook


def move_rows_between_sheets_with_format(file_path, source_sheet, target_sheet, condition_col, condition_val):
    # Load the workbook and the sheets
    workbook = load_workbook(file_path)

    # Load the source and target sheets into pandas DataFrames
    source_df = pd.read_excel(file_path, sheet_name=source_sheet)
    if target_sheet in workbook.sheetnames:
        target_df = pd.read_excel(file_path, sheet_name=target_sheet)
    else:
        target_df = pd.DataFrame()

    # Filter rows from the source sheet based on a condition (e.g., 'Status' == 'Pending')
    rows_to_move = source_df[source_df[condition_col] == condition_val]

    # Append rows to target sheet (my_table_2)
    updated_target_df = pd.concat([target_df, rows_to_move], ignore_index=True)

    # Remove the moved rows from the source sheet (my_table_1)
    remaining_source_df = source_df[source_df[condition_col] != condition_val]

    # Update the sheets with the modified data
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        remaining_source_df.to_excel(writer, sheet_name=source_sheet, index=False)
        updated_target_df.to_excel(writer, sheet_name=target_sheet, index=False)

    # Save the workbook without losing the styles
    workbook.save(file_path)


# Example usage
file_path = r'D:\my_table_1.xlsx'
move_rows_between_sheets_with_format(file_path, source_sheet='my_table_1', target_sheet='my_table_2',
                                     condition_col='Status', condition_val='Pending')
