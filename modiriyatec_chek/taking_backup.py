import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # برای تنظیم رنگ پس‌زمینه
from openpyxl.utils.dataframe import dataframe_to_rows  # وارد کردن تابع dataframe_to_rows

def making_backup(output_file_path, engine):
    # خواندن داده‌ها از جدول به DataFrame
    df = pd.read_sql('SELECT * FROM my_table', con=engine)

    try:
        # بارگذاری فایل اکسل با openpyxl
        workbook = load_workbook(output_file_path)
        sheet = workbook['my_table']  # شیت مورد نظر را انتخاب کنید

        # پاک‌سازی داده‌ها در شیت موجود
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None  # پاک‌سازی سلول‌ها

        # نوشتن داده‌های جدید در شیت موجود
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx + 1, column=c_idx, value=value)  # +1 برای حفظ سرستون‌ها

    except FileNotFoundError:
        # اگر فایل وجود نداشت، فایل جدید ایجاد شود
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='my_table', index=False)

    # تنظیم بک‌گراند ذغالی پس از ذخیره کردن داده‌ها
    # تعریف رنگ ذغالی (خاکستری تیره)
    charcoal_fill = PatternFill(start_color='36454F', end_color='36454F', fill_type='solid')

    # اعمال رنگ ذغالی به تمام سلول‌های دارای داده
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = charcoal_fill

    # ذخیره نهایی فایل
    workbook.save(output_file_path)

    print(f"Data successfully exported to {output_file_path}, with charcoal background.")
